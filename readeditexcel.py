
"""
Created on Tue Mar 23 14:32:22 2021

@author: eceakcicek
"""

# Reading an excel file using Python
import xlrd

# Writing to an excel  
from openpyxl import *

# Detecting keyboard movements such as "Enter"
# import keyboard

# Give the location of the file
#loc = ("path of file")
 
# To open Workbook
#wb = xlrd.open_workbook("denemeexcel.xlsx")
#sheet = wb.sheet_by_index(0)

# To write to the excel
#wd=load_workbook("denemeexcel.xlsx")
#ws=wd["Sheet1"]

# For row 0 and column 0
#print(sheet.cell_value(0, 0))

# Extracting number of rows
#print(sheet.nrows)

# Extracting number of columns
#print(sheet.ncols)

# For rows on column 0
#for i in range(sheet.nrows):
#    print(sheet.cell_value(i, 0))

#(satır, sütun) olacak şekilde
#(1,1) en sol üstteki hücre, (0,0) kabul etmiyor
#wcell1=ws.cell(2,1)
#wcell1.value="Arabanız var mı?"

#wd.save("denemeexcel.xlsx")

wd=load_workbook("denemeexcel.xlsx")
ws=wd["Sheet1"]
i=0
print("press 'o' to see the question, 'c' to correct, 'r' to read the whole file and 'e' to escape")
while True:
    wb = xlrd.open_workbook("denemeexcel.xlsx")
    sheet = wb.sheet_by_index(0)     
    if i==sheet.nrows:
        print("You've reached the EOF. Press 'm' to add new lines!")
    try:
        val= input("Write your command: ")
        if val == "o":
            i=i+1
            print(sheet.cell_value(i-1, 0))
        if val == "c":
            inp= input("Write the corrected question: ")
            wcell1=ws.cell(i,1)
            wcell1.value=inp
            wd.save("denemeexcel.xlsx")
        if val == "a":
            inpt= input("Write the new question: ")
            wcell1=ws.cell(i+1,1)
            wcell1.value=inpt
            wd.save("denemeexcel.xlsx")     
        if val == "r":
            for i in range(sheet.nrows):
                print(sheet.cell_value(i, 0))
        if val == "e":
            break
    except:
        pass